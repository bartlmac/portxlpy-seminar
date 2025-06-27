import openpyxl

from barwerte import act_ngr_ax, act_axn_k
from gwerte import act_dx

from verlaufswerte import (
    calc_Axn,
    calc_axn,
    calc_axt,
    calc_kVx_bpfl,
    calc_kDRx_bpfl,
    calc_kVx_bfr,
    calc_kVx_MRV,
    calc_flex_phase,
    calc_StoAb,
    calc_RKW,
    calc_VS_bfr
)

# Beispielhafte Beitrags-Funktionen 
def calc_Bxt(x, n, t, sex, tafel, zins, alpha, beta1, gamma1, gamma2):
    numerator = (
        act_ngr_ax(x, n, sex, tafel, zins)
        + act_dx(x + n, sex, tafel, zins) / act_dx(x, sex, tafel, zins)
        + gamma1 * act_axn_k(x, t, sex, tafel, zins, 1)
        + gamma2 * (
            act_axn_k(x, n, sex, tafel, zins, 1)
            - act_axn_k(x, t, sex, tafel, zins, 1)
        )
    )
    denominator = (
        (1 - beta1) * act_axn_k(x, t, sex, tafel, zins, 1)
        - alpha * t
    )
    return numerator / denominator

def calc_BJB(VS, Bxt_value):
    return VS * Bxt_value

def calc_BZB(ratzu, zw, BJB_value, k_val):
    return (1 + ratzu) / zw * (BJB_value + k_val)

def calc_Pxt(x, n, t, sex, tafel, zins, alpha, Bxt_value):
    numerator = (
        act_ngr_ax(x, n, sex, tafel, zins)
        + act_dx(x + n, sex, tafel, zins) / act_dx(x, sex, tafel, zins)
        + t * alpha * Bxt_value
    )
    denominator = act_axn_k(x, t, sex, tafel, zins, 1)
    return numerator / denominator

# --------------------------------------------
# Hilfsfunktion zum Toleranz-Vergleich
# --------------------------------------------
DIFF_TOL = 1e-6

def compare_values(label, excel_value, python_value):
    """
    Vergleicht einen Excel-Wert mit dem Python-Ergebnis.
    Gibt das Label, beide Werte und die Differenz aus.
    Kennzeichnet, ob die Differenz innerhalb der Toleranz liegt.
    """
    if excel_value is None:
        print(f"{label:15s}: Excel-Wert ist None (Zelle leer?). Vergleich übersprungen.")
        return
    if python_value is None:
        print(f"{label:15s}: Python-Wert ist None. Vergleich übersprungen.")
        return

    try:
        excel_val_float = float(excel_value)
        python_val_float = float(python_value)
    except ValueError:
        print(f"{label:15s}: Mindestens einer der Werte ist nicht numerisch.\n"
              f"Excel: {excel_value}, Python: {python_value}")
        return

    diff = abs(excel_val_float - python_val_float)
    ok_str = "OK" if diff < DIFF_TOL else "ABWEICHUNG"

    print(f"{label:15s}: Excel={excel_val_float:12.8f}, "
          f"Python={python_val_float:12.8f}, "
          f"Diff={diff:12.8f} => {ok_str}")


def main():
    wb = openpyxl.load_workbook("Tarifrechner_KLV.xlsm", data_only=True)
    sheet = wb["Kalkulation"]

    # Eingabeparameter (Beispiel: Zeilen 4..9 in Spalte B usw.)
    x            = sheet["B4"].value
    sex          = sheet["B5"].value
    n            = sheet["B6"].value
    t            = sheet["B7"].value
    VS           = sheet["B8"].value
    zw           = sheet["B9"].value

    zins         = sheet["E4"].value
    tafel        = sheet["E5"].value
    alpha        = sheet["E6"].value
    beta1        = sheet["E7"].value
    gamma1       = sheet["E8"].value
    gamma2       = sheet["E9"].value
    gamma3       = sheet["E10"].value
    k_val        = sheet["E11"].value
    ratzu        = sheet["E12"].value

    MinAlterFlex = sheet["H4"].value
    MinRLZFlex   = sheet["H5"].value

    # --- Python: Beitragsberechnung ---
    Bxt_value_py = calc_Bxt(x, n, t, sex, tafel, zins, alpha, beta1, gamma1, gamma2)
    BJB_value_py = calc_BJB(VS, Bxt_value_py)
    BZB_value_py = calc_BZB(ratzu, zw, BJB_value_py, k_val)
    Pxt_value_py = calc_Pxt(x, n, t, sex, tafel, zins, alpha, Bxt_value_py)

    # --- Excel: Die Beitragswerte ---
    # laut Anforderung:
    #  Bxt in Zelle K5,
    #  BJB in Zelle K6,
    #  BZB in Zelle K7,
    #  Pxt in Zelle K9
    Bxt_value_xl = sheet["K5"].value
    BJB_value_xl = sheet["K6"].value
    BZB_value_xl = sheet["K7"].value
    Pxt_value_xl = sheet["K9"].value

    print("=== Vergleich: Beitragswerte ===")
    compare_values("Bxt", Bxt_value_xl, Bxt_value_py)
    compare_values("BJB", BJB_value_xl, BJB_value_py)
    compare_values("BZB", BZB_value_xl, BZB_value_py)
    compare_values("Pxt", Pxt_value_xl, Pxt_value_py)
    print()

    print("=== Vergleich: Verlaufswerte (k = 0..n) ===")
    # Verlaufswerte beginnen ab Zeile 16
    for k_ in range(n + 1):
        row_index = 16 + k_

        # Spaltenbelegung anpassen:
        #  Axn  in Spalte B,
        #  axn  in Spalte C,
        #  axt  in Spalte D,
        #  kVx_bpfl  in Spalte E,
        #  kDRx_bpfl in Spalte F,
        #  kVx_bfr   in Spalte G,
        #  kVx_MRV   in Spalte H,
        #  flexPhase in Spalte I,
        #  StoAb     in Spalte J,
        #  RKW       in Spalte K,
        #  VS_bfr    in Spalte L

        Axn_xl       = sheet.cell(row=row_index, column=2).value   # B
        axn_xl       = sheet.cell(row=row_index, column=3).value   # C
        axt_xl       = sheet.cell(row=row_index, column=4).value   # D
        kVx_bpfl_xl  = sheet.cell(row=row_index, column=5).value   # E
        kDRx_bpfl_xl = sheet.cell(row=row_index, column=6).value   # F
        kVx_bfr_xl   = sheet.cell(row=row_index, column=7).value   # G
        kVx_MRV_xl   = sheet.cell(row=row_index, column=8).value   # H
        flex_xl      = sheet.cell(row=row_index, column=9).value   # I
        StoAb_xl     = sheet.cell(row=row_index, column=10).value  # J
        RKW_xl       = sheet.cell(row=row_index, column=11).value  # K
        VS_bfr_xl    = sheet.cell(row=row_index, column=12).value  # L

        # Python-Berechnung
        Axn_py       = calc_Axn(k_, x, n, sex, tafel, zins)
        axn_py       = calc_axn(k_, x, n, sex, tafel, zins)
        axt_py       = calc_axt(k_, x, t, sex, tafel, zins)
        kVx_bpfl_py  = calc_kVx_bpfl(k_, x, n, t, sex, tafel, zins, Pxt_value_py, gamma2)
        kDRx_bpfl_py = calc_kDRx_bpfl(k_, x, n, t, sex, tafel, zins, Pxt_value_py, gamma2, VS)
        kVx_bfr_py   = calc_kVx_bfr(k_, x, n, t, sex, tafel, zins, gamma3)
        kVx_MRV_py   = calc_kVx_MRV(k_, x, n, t, sex, tafel, zins, alpha, BJB_value_py,
                                    Pxt_value_py, gamma2, VS)
        flex_py      = calc_flex_phase(k_, x, n, MinAlterFlex, MinRLZFlex)
        StoAb_py     = calc_StoAb(k_, x, n, t, sex, tafel, zins,
                                  Pxt_value_py, gamma2, VS,
                                  alpha, BJB_value_py,
                                  MinAlterFlex, MinRLZFlex)
        RKW_py       = calc_RKW(k_, x, n, t, sex, tafel, zins,
                                Pxt_value_py, gamma2, VS,
                                alpha, BJB_value_py,
                                MinAlterFlex, MinRLZFlex)
        VS_bfr_py    = calc_VS_bfr(k_, x, n, t, sex, tafel, zins,
                                   Pxt_value_py, gamma2, gamma3, VS,
                                   alpha, BJB_value_py,
                                   MinAlterFlex, MinRLZFlex)

        print(f"\n--- k = {k_} ---")
        compare_values("Axn",        Axn_xl,       Axn_py)
        compare_values("axn",        axn_xl,       axn_py)
        compare_values("axt",        axt_xl,       axt_py)
        compare_values("kVx_bpfl",   kVx_bpfl_xl,  kVx_bpfl_py)
        compare_values("kDRx_bpfl",  kDRx_bpfl_xl, kDRx_bpfl_py)
        compare_values("kVx_bfr",    kVx_bfr_xl,   kVx_bfr_py)
        compare_values("kVx_MRV",    kVx_MRV_xl,   kVx_MRV_py)
        compare_values("flexPhase",  flex_xl,      flex_py)
        compare_values("StoAb",      StoAb_xl,     StoAb_py)
        compare_values("RKW",        RKW_xl,       RKW_py)
        compare_values("VS_bfr",     VS_bfr_xl,    VS_bfr_py)

    wb.close()

if __name__ == "__main__":
    main()
